from aiogram import Router, F, Bot
from aiogram.types import Message, FSInputFile
from aiogram.enums.content_type import ContentType
from openpyxl import Workbook, load_workbook
from telethon import TelegramClient, functions
from telethon.errors.rpcerrorlist import FloodWaitError
import asyncio
import os


router = Router()


@router.message(F.content_type.in_({ContentType.DOCUMENT}))
async def check_file_type(message: Message, bot: Bot, client: TelegramClient):
    SLEEP = 2
    UPDATE_METHOD = 5
    SLEEP_TABLE = 1
    UPDATE_TABLE = 10
    if message.document.file_name.split(".")[-1] == "xlsx":
        bot_message = await message.answer("✅")

        file = await bot.get_file(message.document.file_id)
        file_path = file.file_path
        await bot.download_file(file_path, f"tmp\\{message.chat.id}.xlsx")
        
        wb = load_workbook(f"tmp\\{message.chat.id}.xlsx")
        sheet = wb[wb.sheetnames[0]]
        urls = []
        updater = UPDATE_TABLE
        for i, row in enumerate(sheet.values):
            if updater == UPDATE_TABLE:
                updater = 0
                await bot_message.edit_text(f"🔍 Сбор ссылок из <b><i>{message.document.file_name}</i></b> (~{SLEEP_TABLE * int((len([j for j in sheet.values]) - i) / UPDATE_TABLE)} с.):\n<b>{i+1} / {len([j for j in sheet.values])}</b>")
                await asyncio.sleep(SLEEP_TABLE)
            updater += 1
            url = row[0]
            if url is None:
                continue
            if "+" in url or "joinchat" in url:
                continue
            else:
                urls.append("@" + url.split("/")[-1])
        wb.save(f"tmp\\{message.chat.id}.xlsx")

        usernames = []
        time_work = round(len(urls)*SLEEP/60)
        async with client:
            updater = UPDATE_METHOD
            for url in urls:
                if updater == UPDATE_METHOD:
                    updater = 0
                    await bot_message.edit_text(f"📢 Получение ссылок похожих каналов (<b><i>~{time_work - int(urls.index(url) * SLEEP / 60)} мин.</i></b>):\n<b>{urls.index(url)+1} / {len(urls)}</b>")
                updater += 1
                await asyncio.sleep(SLEEP)
                try:
                    result = await client(functions.channels.GetChannelRecommendationsRequest(
                        channel=url
                    ))
                    for channel in result.chats:
                        if channel.username is None:
                            usernames.append("@" + channel.usernames[0].username)
                        else:
                            usernames.append("@" + channel.username)
                except FloodWaitError as error:
                    print(error)
                    await asyncio.sleep(60)
        usernames = list(set(usernames))
        channels = []
        updater = UPDATE_METHOD
        time_work = round(len(usernames)*SLEEP/60)
        await asyncio.sleep(SLEEP)
        async with client:
            for username in usernames:
                if updater == UPDATE_METHOD:
                    updater = 0
                    await bot_message.edit_text(f"📲 Сохранение информации о похожих каналах (<b><i>~{time_work - int(usernames.index(username) * SLEEP / 60)} мин.</i></b>):\n<b>{usernames.index(username)+1} / {len(usernames)}</b>")
                updater += 1
                await asyncio.sleep(SLEEP)
                try:
                    channel = await client(functions.channels.GetFullChannelRequest(
                        channel=username
                    ))
                    title = channel.chats[0].title
                    about = channel.full_chat.about
                    link = f"https://t.me/{username[1:]}"
                    subscibers = channel.full_chat.participants_count
                    channels.append([title, about, link, subscibers])
                except FloodWaitError as error:
                    print(error)
                    await asyncio.sleep(60)

        wb = Workbook()
        sheet = wb[wb.sheetnames[0]]
        updater = UPDATE_TABLE
        for row, value in enumerate(channels):
            if updater == UPDATE_TABLE:
                updater = 0
                await bot_message.edit_text(f"📝 Загрузка каналов в <i><b>similar_channels.xlsx</b></i> (~{SLEEP_TABLE * ((len(channels) - row) / UPDATE_TABLE)} с.):\n<b>{row+1} / {len(channels)}</b>")
            updater += 1
            for column, v in enumerate(value):
                sheet.cell(row=row+1, column=column+1, value=v)
        wb.save(f"tmp\\new_{message.chat.id}.xlsx")

        await bot_message.delete()
        file = FSInputFile(f"tmp\\new_{message.chat.id}.xlsx", filename="similar_channels.xlsx")
        await message.answer_document(file)

        os.remove(f"tmp\\{message.chat.id}.xlsx")
        os.remove(f"tmp\\new_{message.chat.id}.xlsx")

    else:
        await message.answer("‼️ Вы отправили не .xlsx файл!\nОтправьте нужный файл.")